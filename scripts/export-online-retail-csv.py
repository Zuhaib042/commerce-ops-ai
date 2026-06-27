from __future__ import annotations

import csv
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "Missing dependency: openpyxl. Install it with "
        "`npm run py:venv && npm run py:install`."
    )


ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw" / "online-retail"
OUTPUT_PATH = RAW_DIR / "online_retail.csv"


def main() -> None:
    workbook_paths = sorted(RAW_DIR.glob("*.xlsx"))
    if not workbook_paths:
        sys.exit("No .xlsx workbook found. Run `npm run data:download` first.")

    workbook_path = workbook_paths[0]
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    with OUTPUT_PATH.open("w", newline="", encoding="utf-8") as output_file:
        writer = csv.writer(output_file)
        for row in worksheet.iter_rows(values_only=True):
            writer.writerow(row)

    workbook.close()
    print(f"Exported {workbook_path} to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
